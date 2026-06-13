"""Import Excel, normalisation des colonnes et conversion des données.

Le chargeur lit dynamiquement le modèle Excel OptiFLUX, sans hypothèse sur le
nombre de lignes. Les noms métier restent dans le fichier source ; le code ne
contient que les colonnes obligatoires et leurs alias techniques.
"""

from __future__ import annotations

import math
import re
import unicodedata
from datetime import datetime, time
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from config import (
    COL_ALIASES,
    DEFAULT_DOCK_CAPACITY,
    EXPECTED_SHEETS,
    NC_VALUES,
    QUANTITY_PREFIX,
    REQUIRED_COLUMNS,
    SHEET_CONTAINERS,
    SHEET_DISTANCE,
    SHEET_DURATION,
    SHEET_FLOWS,
    SHEET_RH,
    SHEET_SITES,
    SHEET_VEHICLES,
    WEEKDAYS,
    YES_VALUES,
)
from exceptions import Diagnostic, ImportBlockingError
from models import ContainerType, RHParams, Site, VehicleType


def _safe_str(val: Any) -> str | None:
    """Retourne une chaîne nettoyée ou None pour les valeurs vides/NaN/'nan'."""
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    s = str(val).strip()
    if not s or s.casefold() in {"nan", "none", "null"}:
        return None
    return s


def _ns(text: Any) -> str:
    """Normalise un libellé pour les comparaisons de colonnes."""
    s = "" if text is None else str(text)
    s = s.replace("\n", " ").replace("\r", " ")
    s = " ".join(s.strip().split())
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.casefold()


def _to_bool(val: Any, default: bool = False) -> bool:
    s = _safe_str(val)
    if s is None:
        return default
    return _ns(s) in YES_VALUES


def _to_minutes(value: Any, *, default: int | None = None, field: str = "horaire") -> int:
    """Convertit un horaire Excel/time/HH:MM en minutes depuis minuit.

    Les durées de manutention dans le modèle sont souvent stockées en fraction de
    journée ; la même conversion est donc utilisée pour les durées et horaires.
    """
    if value is None or (isinstance(value, float) and math.isnan(value)):
        if default is not None:
            return default
        raise ValueError(f"Valeur horaire absente pour {field}")
    if isinstance(value, time):
        return int(round(value.hour * 60 + value.minute + value.second / 60))
    if isinstance(value, datetime):
        return int(round(value.hour * 60 + value.minute + value.second / 60))
    if isinstance(value, pd.Timestamp):
        return int(round(value.hour * 60 + value.minute + value.second / 60))
    if isinstance(value, (int, float)):
        if 0 <= float(value) <= 1:
            return int(round(float(value) * 1440))
        return int(round(float(value)))
    s = _safe_str(value)
    if s is None:
        if default is not None:
            return default
        raise ValueError(f"Valeur horaire absente pour {field}")
    if _ns(s) in NC_VALUES:
        raise ValueError(f"NC n'est pas un horaire pour {field}")
    m = re.match(r"^(\d{1,2})[:hH](\d{2})$", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        x = float(s.replace(",", "."))
        return _to_minutes(x, default=default, field=field)
    except ValueError as exc:
        raise ValueError(f"Format horaire non reconnu pour {field}: {s}") from exc


def _to_optional_duration_minutes(value: Any) -> float | None:
    s = _safe_str(value)
    if s is None or _ns(s) in NC_VALUES:
        return None
    return float(_to_minutes(value, default=0, field="durée"))


def _num(value: Any, default: float = 0.0) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if not s:
            return default
        return float(s)
    return float(value)


def _quantity_to_int(value: Any) -> int:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0
    if isinstance(value, str) and value.strip().startswith("="):
        raise ValueError("Formule Excel non calculée")
    return int(round(_num(value, 0.0)))


def _read_excel(path: str | Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, engine="openpyxl")


def _resolve_columns(df: pd.DataFrame, sheet_name: str) -> dict[str, str]:
    normalized = {_ns(c): c for c in df.columns}
    mapping: dict[str, str] = {}
    for canonical, aliases in COL_ALIASES.items():
        for alias in aliases:
            if _ns(alias) in normalized:
                mapping[canonical] = normalized[_ns(alias)]
                break
    return mapping


def _missing_required_columns(df: pd.DataFrame, sheet_name: str) -> list[Diagnostic]:
    mapping = _resolve_columns(df, sheet_name)
    diagnostics: list[Diagnostic] = []
    for canonical in REQUIRED_COLUMNS.get(sheet_name, []):
        if canonical not in mapping:
            diagnostics.append(
                Diagnostic(
                    sheet=sheet_name,
                    row=1,
                    column=None,
                    severity="CRITIQUE",
                    code="COLONNE_MANQUANTE",
                    message=f"Colonne obligatoire introuvable : {canonical}. Alias acceptés : {COL_ALIASES.get(canonical, [])}",
                    action="Vérifier l'en-tête de colonne ou ajouter un alias dans config.py.",
                )
            )
    return diagnostics


def detect_formula_quantities(path: str | Path) -> list[Diagnostic]:
    """Détecte les formules dans les colonnes Quantité Lundi-Dimanche de M flux."""
    diagnostics: list[Diagnostic] = []
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    if SHEET_FLOWS not in wb.sheetnames:
        return diagnostics
    ws = wb[SHEET_FLOWS]
    headers = [cell.value for cell in ws[1]]
    qty_cols = []
    for idx, header in enumerate(headers, start=1):
        if _safe_str(header) and _ns(header).startswith(_ns(QUANTITY_PREFIX)):
            qty_cols.append((idx, str(header)))
    for row in range(2, ws.max_row + 1):
        for col, header in qty_cols:
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val.strip().startswith("="):
                diagnostics.append(
                    Diagnostic(
                        sheet=SHEET_FLOWS,
                        row=row,
                        column=openpyxl.utils.get_column_letter(col),
                        severity="CRITIQUE",
                        code="FORMULE_QUANTITE",
                        message=f"Formule Excel non calculée dans {header} : {val}",
                        action="Remplacer la formule par sa valeur ou enregistrer le fichier avec calcul des formules.",
                    )
                )
    return diagnostics


class OptiFluxData:
    """Conteneur simple des données importées."""

    def __init__(self, **kwargs: Any) -> None:
        self.__dict__.update(kwargs)


def load_workbook_data(path: str | Path) -> OptiFluxData:
    """Charge et valide structurellement le fichier Excel source."""
    path = Path(path)
    diagnostics: list[Diagnostic] = []
    try:
        sheets = _read_excel(path)
    except Exception as exc:  # pragma: no cover - protection UI
        raise ImportBlockingError([
            Diagnostic(None, None, None, "CRITIQUE", "FICHIER_ILLISIBLE", f"Impossible de lire le fichier Excel : {exc}")
        ]) from exc

    missing_sheets = [s for s in EXPECTED_SHEETS if s not in sheets]
    for sheet in missing_sheets:
        diagnostics.append(Diagnostic(sheet, None, None, "CRITIQUE", "ONGLET_MANQUANT", f"Onglet attendu absent : {sheet}"))
    if missing_sheets:
        raise ImportBlockingError(diagnostics)

    diagnostics.extend(detect_formula_quantities(path))
    for sheet_name in REQUIRED_COLUMNS:
        diagnostics.extend(_missing_required_columns(sheets[sheet_name], sheet_name))

    if diagnostics:
        raise ImportBlockingError(diagnostics)

    column_maps = {sheet: _resolve_columns(df, sheet) for sheet, df in sheets.items()}

    rh = parse_rh(sheets[SHEET_RH], column_maps[SHEET_RH])
    sites = parse_sites(sheets[SHEET_SITES], column_maps[SHEET_SITES])
    containers = parse_containers(sheets[SHEET_CONTAINERS], column_maps[SHEET_CONTAINERS])
    vehicles = parse_vehicles(sheets[SHEET_VEHICLES], column_maps[SHEET_VEHICLES], containers)
    duration = parse_matrix(sheets[SHEET_DURATION])
    distance = parse_matrix(sheets[SHEET_DISTANCE])

    return OptiFluxData(
        path=path,
        raw=sheets,
        column_maps=column_maps,
        rh=rh,
        sites=sites,
        containers=containers,
        vehicles=vehicles,
        duration_matrix=duration,
        distance_matrix=distance,
        import_warnings=[],
    )


def parse_rh(df: pd.DataFrame, cols: dict[str, str]) -> RHParams:
    row = df.dropna(how="all").iloc[0]
    vacation = _to_minutes(row[cols["rh_vacation"]], field="Durée de vacation")
    pause = _to_minutes(row[cols["rh_pause"]], field="Pause")
    start_min = _to_minutes(row[cols["rh_start_min"]], field="Heure début mini")
    end_max = _to_minutes(row[cols["rh_end_max"]], field="Heure fin max")
    return RHParams(vacation_duration=vacation, pause_duration=pause, start_min=start_min, end_max=end_max)


def parse_sites(df: pd.DataFrame, cols: dict[str, str]) -> dict[str, Site]:
    vehicle_cols = [c for c in df.columns if c not in {cols.get("site_name"), cols.get("site_address"), cols.get("site_has_dock"), cols.get("site_dock_capacity")}]
    sites: dict[str, Site] = {}
    for _, row in df.iterrows():
        name = _safe_str(row.get(cols["site_name"]))
        if name is None:
            continue
        capacity = DEFAULT_DOCK_CAPACITY
        if "site_dock_capacity" in cols:
            raw_capacity = row.get(cols["site_dock_capacity"])
            if raw_capacity is not None and not (isinstance(raw_capacity, float) and math.isnan(raw_capacity)):
                capacity = max(1, int(round(_num(raw_capacity, DEFAULT_DOCK_CAPACITY))))
        compat = {str(vcol): _to_bool(row.get(vcol), default=False) for vcol in vehicle_cols if _safe_str(vcol)}
        sites[name] = Site(
            name=name,
            address=_safe_str(row.get(cols.get("site_address"))),
            has_dock=_to_bool(row.get(cols["site_has_dock"]), default=True),
            dock_capacity=capacity,
            vehicle_compat=compat,
        )
    return sites


def parse_containers(df: pd.DataFrame, cols: dict[str, str]) -> dict[str, ContainerType]:
    containers: dict[str, ContainerType] = {}
    for _, row in df.iterrows():
        name = _safe_str(row.get(cols["container_name"]))
        if name is None:
            continue
        containers[name] = ContainerType(
            name=name,
            length_m=_num(row.get(cols["container_length"])),
            width_m=_num(row.get(cols["container_width"])),
            empty_weight_t=_num(row.get(cols["container_empty_weight"])),
            full_weight_t=_num(row.get(cols["container_full_weight"])),
        )
    return containers


def parse_vehicles(df: pd.DataFrame, cols: dict[str, str], containers: dict[str, ContainerType]) -> dict[str, VehicleType]:
    fixed_cols = {cols[k] for k in cols if k.startswith("vehicle_") and k in cols}
    container_cols = [c for c in df.columns if c not in fixed_cols and _safe_str(c)]
    vehicles: dict[str, VehicleType] = {}
    for _, row in df.iterrows():
        vtype = _safe_str(row.get(cols["vehicle_type"]))
        if vtype is None:
            continue
        compat = {str(c): _to_bool(row.get(c), default=False) for c in container_cols}
        # Garantit que les contenants absents de la matrice de compatibilité sont considérés incompatibles.
        for c_name in containers:
            compat.setdefault(c_name, False)
        vehicles[vtype] = VehicleType(
            type=vtype,
            initial_site=_safe_str(row.get(cols["vehicle_initial_site"])) or "",
            length_m=_num(row.get(cols["vehicle_length"])),
            width_m=_num(row.get(cols["vehicle_width"])),
            height_m=_num(row.get(cols["vehicle_height"]), 0.0),
            max_weight_t=_num(row.get(cols["vehicle_max_weight"])),
            consumption_l_km=_num(row.get(cols["vehicle_consumption"]), 0.0),
            fuel_cost_per_km=_num(row.get(cols["vehicle_fuel_cost"]), 0.0),
            co2_kg_per_km=_num(row.get(cols["vehicle_co2"]), 0.0),
            has_tail_lift=_to_bool(row.get(cols["vehicle_tail_lift"]), default=False),
            container_compat=compat,
            dock_time_min=_to_minutes(row.get(cols["vehicle_dock_time"]), default=0, field="Temps de mise à quai"),
            manual_no_dock_min_per_container=_to_optional_duration_minutes(row.get(cols["vehicle_manual_no_dock"])),
            manual_dock_min_per_container=float(_to_minutes(row.get(cols["vehicle_manual_dock"]), default=0, field="Manutention avec quai")),
        )
    return vehicles


def parse_matrix(df: pd.DataFrame) -> dict[tuple[str, str], float]:
    """Parse une matrice carrée où A1 est vide, ligne 1 = destinations et colonne A = origines."""
    if df.empty:
        return {}
    df2 = df.copy()
    first_col = df2.columns[0]
    matrix: dict[tuple[str, str], float] = {}
    destinations = [_safe_str(c) for c in df2.columns[1:]]
    for _, row in df2.iterrows():
        origin = _safe_str(row.get(first_col))
        if origin is None:
            continue
        for col, dest in zip(df2.columns[1:], destinations):
            if dest is None:
                continue
            matrix[(origin, dest)] = _num(row.get(col), 0.0)
    return matrix


def get_flow_dataframe(data: OptiFluxData) -> pd.DataFrame:
    return data.raw[SHEET_FLOWS].copy()


def get_quantity_column(df: pd.DataFrame, day: str) -> str:
    candidates = [c for c in df.columns if _ns(c) == _ns(f"{QUANTITY_PREFIX} {day}")]
    if not candidates:
        raise KeyError(f"Colonne Quantité {day} introuvable")
    return candidates[0]
